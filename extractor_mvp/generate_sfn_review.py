#!/usr/bin/env python3
"""
SfN v5 batch review spreadsheet generator.

Usage (from extractor_mvp/):
    python generate_sfn_review.py
    python generate_sfn_review.py --results-dir results/batch_sfn_v5 --output sfn_review_v5.xlsx

Sheets:
    1. Priority Review  — EXTRACTED + out-of-vocab + quote-unresolved rows
    2. target_space     — All 20 papers x target_space (the 10/20 headline finding)
    3. All Fields       — Complete record
    4. Glossary         — All step_kinds, fields, valid values, descriptions
"""

import argparse
import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

COLOURS = {
    "Extracted ✓": "C6EFCE",
    "Out-of-vocab ⚠": "FFEB9C",
    "Quote unresolved": "FFD7B5",
    "Deferred ↗": "DDEEFF",
    "Deferral unresolved": "FFD7B5",
    "Missing": "FFFFFF",
}

GLOSSARY = [
    (
        "nonsteadystate_removal",
        "n_nonsteadystate_discarded",
        "Dummy scans discarded for T1 signal stabilization at run start",
        "integer ≥ 0",
        "",
    ),
    (
        "slice_time_correction",
        "reference",
        "Reference time point for inter-slice timing alignment",
        "first | middle | specific_slice | specific_time",
        "",
    ),
    (
        "slice_time_correction",
        "relative_to_motion_correction",
        "Whether STC was applied before or after motion correction",
        "before | after",
        "",
    ),
    (
        "slice_time_correction",
        "interpolation",
        "Temporal interpolation method",
        "linear | spline | sinc",
        "",
    ),
    (
        "motion_correction",
        "method",
        "Software tool used for head-motion realignment",
        "mcflirt | spm_realign | afni_3dvolreg | ants | other",
        "",
    ),
    (
        "motion_correction",
        "reference_scan",
        "Reference volume each frame is registered to",
        "first | middle | mean | specific",
        "",
    ),
    (
        "motion_correction",
        "similarity_metric",
        "Cost function optimized during realignment",
        "normalized_correlation | mutual_information | ssd | other",
        "",
    ),
    (
        "motion_correction",
        "interpolation",
        "Resampling method applied during motion correction",
        "linear | spline | sinc | other",
        "",
    ),
    (
        "motion_correction",
        "nonrigid",
        "Whether non-rigid-body correction was applied",
        "true | false",
        "",
    ),
    (
        "motion_correction",
        "transform_type",
        "Transform degrees of freedom (e.g. '6dof rigid')",
        "free text",
        "",
    ),
    (
        "motion_correction",
        "fieldmap_unwarping",
        "Whether B0 distortion was corrected jointly with motion",
        "true | false",
        "",
    ),
    (
        "motion_correction",
        "unwarping_method",
        "Unwarping approach when fieldmap_unwarping=True",
        "free text",
        "",
    ),
    (
        "motion_correction",
        "slice_to_volume",
        "Whether slice-to-volume correction was applied",
        "true | false",
        "",
    ),
    (
        "distortion_correction",
        "source",
        "Source of the distortion being corrected",
        "susceptibility_fieldmap | gradient_nonlinearity | fieldmap_less",
        "",
    ),
    (
        "distortion_correction",
        "method",
        "Tool used for distortion correction",
        "topup | fugue | gradunwarp | sdc_fieldmapless | other",
        "",
    ),
    (
        "coregistration",
        "transform",
        "Transform type for functional↔structural alignment",
        "rigid | affine | nonlinear",
        "",
    ),
    (
        "coregistration",
        "method",
        "Tool for intra-subject coregistration",
        "flirt_bbr | flirt | spm_coreg | bbregister | ants | other",
        "",
    ),
    (
        "coregistration",
        "cost_function",
        "Optimization criterion",
        "correlation_ratio | mutual_information | boundary_based | ssd",
        "",
    ),
    (
        "coregistration",
        "interpolation",
        "Resampling method at coregistration",
        "linear | spline | sinc | other",
        "",
    ),
    ("intensity_correction", "target", "What was corrected", "bias_field | interleaved_slice", ""),
    (
        "intensity_correction",
        "method",
        "Tool for intensity correction",
        "n4 | fast_bias | other",
        "",
    ),
    (
        "spatial_normalization",
        "target_space",
        "MNI space variant. Out-of-vocab = paper said 'MNI'/'MNI152' without the NLin variant — this is the 10/20 headline finding.",
        "MNI152NLin6Asym | MNI152NLin2009cAsym | Talairach | native_volume | other",
        "",
    ),
    (
        "spatial_normalization",
        "resolution_mm",
        "Isotropic voxel size IN ATLAS SPACE after normalization. NOT acquisition voxel size; NOT ROI sphere radius.",
        "float",
        "mm",
    ),
    (
        "spatial_normalization",
        "method",
        "Normalization tool",
        "fnirt | ants_syn | spm_normalise | dartel | other",
        "",
    ),
    ("spatial_normalization", "warp", "Transform complexity", "rigid | affine | nonlinear", ""),
    ("spatial_normalization", "transform_type", "Additional transform details", "free text", ""),
    (
        "spatial_normalization",
        "interpolation",
        "Final resampling interpolation",
        "linear | spline | sinc | other",
        "",
    ),
    ("spatial_normalization", "regularization", "Regularization approach", "free text", ""),
    (
        "surface_projection",
        "target_surface",
        "Surface template to which volume data were projected",
        "native | fsaverage | fsaverage5 | fsaverage6 | fsLR_32k | fsLR_164k | other",
        "",
    ),
    (
        "surface_projection",
        "vol2surf_sampling",
        "Method for sampling volume data onto surface",
        "ribbon_constrained | trilinear | nearest",
        "",
    ),
    (
        "surface_projection",
        "surface_registration",
        "Surface registration approach",
        "freesurfer_recon | msm_sulc | msm_all | other",
        "",
    ),
    ("surface_projection", "cifti", "Whether output is CIFTI format", "true | false", ""),
    ("ica_denoise", "method", "ICA-based denoising approach", "fix | aroma", ""),
    ("ica_denoise", "training_set", "Training dataset for classifier", "free text", ""),
    ("ica_denoise", "threshold", "Decision threshold for component classification", "float", ""),
    ("ica_denoise", "aggressive", "Aggressive vs non-aggressive denoising", "true | false", ""),
    ("compcor", "variant", "aCompCor (anatomical) vs tCompCor (temporal)", "a | t", ""),
    ("compcor", "n_components", "Number of CompCor components used as regressors", "integer", ""),
    (
        "compcor",
        "variance_threshold",
        "Cumulative variance threshold for component selection",
        "float 0-1",
        "",
    ),
    ("compcor", "mask_source", "Tissue mask for CompCor ROI", "free text", ""),
    (
        "nuisance_regression",
        "motion_expansion",
        "Motion regressor set",
        "none | 6param | friston24 | volterra",
        "",
    ),
    (
        "nuisance_regression",
        "tissue_regressors",
        "Tissue-compartment signal regressors (list)",
        "whole_brain | gray_matter | white_matter | ventricles",
        "",
    ),
    (
        "nuisance_regression",
        "physio_regressors",
        "Physiological noise modelling",
        "retroicor | rvt | none",
        "",
    ),
    ("nuisance_regression", "physio_n_regressors", "Number of RETROICOR regressors", "integer", ""),
    ("nuisance_regression", "detrend", "Polynomial detrending", "linear | quadratic | none", ""),
    ("despike", "method", "Despiking tool", "afni_3dDespike | other", ""),
    ("despike", "threshold", "Spike detection threshold", "float", ""),
    (
        "scrub",
        "criterion",
        "Motion metric for censoring",
        "fd_power | fd_jenkinson | dvars | bold_pct",
        "",
    ),
    ("scrub", "threshold", "Censoring threshold value", "float", ""),
    ("scrub", "remediation", "How flagged volumes are handled", "censor | interpolate", ""),
    (
        "scrub",
        "interpolation_method",
        "Interpolation method for censored volumes",
        "spline | spectral | other",
        "",
    ),
    (
        "temporal_filtering",
        "effective_band_hz",
        "Canonical passband [low_Hz, high_Hz]",
        "(float|None, float|None) tuple",
        "Hz",
    ),
    (
        "temporal_filtering",
        "method",
        "Filter implementation",
        "butterworth_bandpass | highpass_only | wavelet_decomposition",
        "",
    ),
    ("temporal_filtering", "low_hz", "Lower cutoff (Butterworth)", "float", "Hz"),
    ("temporal_filtering", "high_hz", "Upper cutoff (Butterworth)", "float", "Hz"),
    ("temporal_filtering", "order", "Filter order", "integer", ""),
    ("temporal_filtering", "cutoff", "Cutoff for highpass-only", "float", "Hz"),
    ("temporal_filtering", "scale", "Wavelet scale index", "integer", ""),
    (
        "temporal_filtering",
        "nominal_band_hz",
        "Wavelet scale nominal frequency range",
        "(float, float) tuple",
        "Hz",
    ),
    ("intensity_normalization", "scope", "Spatial scope of normalization", "free text", ""),
    (
        "intensity_normalization",
        "convention",
        "Normalization approach",
        "voxel_temporal_zscore | global_median_1000 | global_mode_1000 | fsl_grand_mean_10000 | fsl_median_10000",
        "",
    ),
    (
        "intensity_normalization",
        "value",
        "Target intensity value (e.g. 1000, 10000). None for z-score.",
        "float | None",
        "",
    ),
    ("spatial_smoothing", "fwhm_mm", "Smoothing kernel FWHM", "float", "mm"),
    ("spatial_smoothing", "space", "Domain of smoothing", "volume | surface | both", ""),
    ("spatial_smoothing", "kernel_type", "Kernel shape", "gaussian | other", ""),
    (
        "spatial_smoothing",
        "approach",
        "Smoothing implementation",
        "fsl_susan | afni_3dBlurInMask | fslmaths | other",
        "",
    ),
]

STATUS_LEGEND = [
    (
        "Extracted ✓",
        "C6EFCE",
        "Pipeline found a value with a verified character span. Verify value + quote match the paper.",
    ),
    (
        "Out-of-vocab ⚠",
        "FFEB9C",
        "LLM extracted something but it didn't match any canonical Literal. pipeline_value shows the raw extracted value. This is the 10/20 finding.",
    ),
    (
        "Quote unresolved",
        "FFD7B5",
        "Pipeline returned a value but span resolver couldn't locate the quote. pipeline_value shows the raw value; verbatim_quote shows what the LLM returned.",
    ),
    (
        "Deferred ↗",
        "DDEEFF",
        "Paper explicitly delegates to a cited paper for preprocessing details.",
    ),
    ("Deferral unresolved", "FFD7B5", "Paper defers but deferral sentence couldn't be located."),
    ("Missing", "FFFFFF", "Field not stated in the methods section."),
]


def classify(extraction: dict, inference: dict) -> tuple[str, str, str]:
    ext = extraction.get("status", "")
    inf_reason = ""
    if inference.get("status") not in ("NOT_APPLICABLE", None):
        inf_reason = inference.get("reason", "")

    if ext == "EXTRACTED":
        value = str(extraction.get("value", ""))
        spans = extraction.get("spans", [])
        quote = spans[0].get("text", "")[:400].replace("\n", " ") if spans else ""
        return "Extracted ✓", value, quote

    if ext == "DEFERRED_TO_CITATION":
        deferrals = extraction.get("deferrals", [])
        ref = deferrals[0].get("ref", "") if deferrals else ""
        return "Deferred ↗", "", ref

    reason_map = {
        "value_not_in_literal": "Out-of-vocab ⚠",
        "extraction_quote_unresolved": "Quote unresolved",
        "deferral_quote_unresolved": "Deferral unresolved",
        "deferred_pending_citation_resolution": "Deferred (pending)",
    }
    return reason_map.get(inf_reason, "Missing"), "", ""


def build_df(results_dir: Path) -> pd.DataFrame:
    rows = []
    papers_dir = results_dir / "papers"
    if not papers_dir.exists():
        raise SystemExit(f"Not found: {papers_dir}")

    for json_path in sorted(papers_dir.glob("*.json")):
        data = json.loads(json_path.read_text())
        paper_id = data["paper_id"]

        # Build diagnostics lookup: "step_kind.field" → diagnostic dict
        diag_lookup: dict[str, dict] = {}
        for diag in data.get("diagnostics", []):
            field_path = diag.get("field", "")
            diag_lookup[field_path] = diag

        for step in data.get("preprocessing", {}).get("steps", []):
            kind = step.get("kind", "")
            for field, pf in step.items():
                if field == "kind" or not isinstance(pf, dict) or "extraction" not in pf:
                    continue

                status, value, quote = classify(pf["extraction"], pf.get("inference", {}))

                # For non-extracted rows: fill pipeline_value and verbatim_quote
                # from diagnostics so reviewers can see what the LLM actually found
                if status != "Extracted ✓":
                    field_path = f"{kind}.{field}"
                    diag = diag_lookup.get(field_path, {})
                    if diag.get("raw_value") is not None and not value:
                        value = str(diag["raw_value"])
                    if diag.get("raw_quote") and not quote:
                        quote = diag["raw_quote"][:400].replace("\n", " ")

                rows.append(
                    {
                        "paper_id": paper_id,
                        "step_kind": kind,
                        "field": field,
                        "status": status,
                        "pipeline_value": value,
                        "verbatim_quote": quote,
                        "review": "",
                        "correction": "",
                        "notes": "",
                    }
                )
    return pd.DataFrame(rows)


def style_data_sheet(ws, df_sheet: pd.DataFrame):
    hdr_fill = PatternFill("solid", start_color="2F5496")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_letter, width in zip("ABCDEFGHI", [18, 22, 22, 20, 18, 55, 10, 20, 30], strict=False):
        ws.column_dimensions[col_letter].width = width

    for i, row in enumerate(ws.iter_rows(min_row=2)):
        if i >= len(df_sheet):
            break
        hex_col = COLOURS.get(df_sheet.iloc[i]["status"], "FFFFFF")
        fill = PatternFill("solid", start_color=hex_col)
        for cell in row:
            cell.fill = fill
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30


def write_glossary_sheet(ws):
    ws.append(["STATUS LEGEND", "", "", "", ""])
    ws.append(["Status", "Colour", "Meaning", "", ""])
    for status, _colour, meaning in STATUS_LEGEND:
        ws.append([status, "", meaning, "", ""])

    ws.append(["", "", "", "", ""])
    ws.append(["FIELD GLOSSARY", "", "", "", ""])
    ws.append(["step_kind", "field", "description", "valid_values", "unit"])
    for row in GLOSSARY:
        ws.append(list(row))

    title_font = Font(bold=True, name="Arial", size=11)
    hdr_fill = PatternFill("solid", start_color="2F5496")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_num in [1, 5]:
        ws.cell(row=row_num, column=1).font = title_font
    for row_num in [2, 6]:
        for cell in ws[row_num]:
            cell.fill = hdr_fill
            cell.font = hdr_font

    for i, (_, colour, _) in enumerate(STATUS_LEGEND, start=3):
        ws.cell(row=i, column=2).fill = PatternFill("solid", start_color=colour)
        ws.cell(row=i, column=2).value = "  "

    for col_letter, width in zip("ABCDE", [28, 28, 65, 55, 8], strict=False):
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A7"


def write_excel(df: pd.DataFrame, output: Path):
    priority_mask = df["status"].isin(["Extracted ✓", "Out-of-vocab ⚠", "Quote unresolved"])
    ts_mask = df["field"] == "target_space"

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        data_sheets = [
            ("1 Priority Review", df[priority_mask].reset_index(drop=True)),
            ("2 target_space", df[ts_mask].reset_index(drop=True)),
            ("3 All Fields", df.reset_index(drop=True)),
        ]
        for name, df_s in data_sheets:
            df_s.to_excel(writer, sheet_name=name, index=False)
            style_data_sheet(writer.sheets[name], df_s)

        writer.book.create_sheet("4 Glossary")
        write_glossary_sheet(writer.book["4 Glossary"])

    print(f"\nWritten: {output}")
    for name, df_s in data_sheets:
        print(f"  {name}: {len(df_s)} rows")
    print(f"  4 Glossary: {len(GLOSSARY)} fields across 16 step kinds")
    print("\nReview: y (agree)  n (disagree — fill correction)  ? (uncertain)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--results-dir", default="results/batch_sfn_v5")
    ap.add_argument("--output", default="sfn_review_v5.xlsx")
    args = ap.parse_args()
    write_excel(build_df(Path(args.results_dir)), Path(args.output))


if __name__ == "__main__":
    main()
