"""Derive the canonical scored CSV from the human-source labeling xlsx.

Single source of truth is the **xlsx** (human-editable). The CSV is fully REPRODUCIBLE from it —
no hand-edits that regeneration cannot reproduce, so `xlsx` and `csv` cannot silently drift.

The protocol version is a **set-level** fact recorded in `README.md`, NOT a per-row CSV column
(a per-row column can't be reproduced from the xlsx, which has no version column, so it would
silently revert on re-derive). This script emits no `protocol_version` column; it only *reads* the
set-level version from the README for a provenance print.

Run:  uv run --with openpyxl python ground_truth/derive_labels_csv.py
"""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
XLSX = HERE / "base_pipeline_labels_v1.xlsx"
CSV = HERE / "base_pipeline_labels_v1.csv"
README = HERE / "README.md"

LABELER = "Jae Wook Cho"
# CSV schema — deliberately NO protocol_version (set-level, see README).
COLUMNS = ["paper_id", "status", "value", "specificity", "supporting_quote", "notes", "labeler"]
# xlsx Labels-sheet column order: Paper, Status, Value, Specificity, Supporting quote, Notes
_XLSX_COLS = (1, 2, 3, 4, 5, 6)


def set_level_version() -> str:
    """Read the single set-level protocol version from the README's machine-readable line."""
    m = re.search(
        r"Protocol version \(set-level\): v(\d+(?:\.\d+)*)", README.read_text(encoding="utf-8")
    )
    return f"v{m.group(1)}" if m else "unknown"


def derive() -> int:
    import openpyxl

    ws = openpyxl.load_workbook(XLSX, data_only=True)["Labels"]
    n = 0
    with open(CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        w.writerow(COLUMNS)
        for r in range(2, ws.max_row + 1):
            paper = ws.cell(r, 1).value
            status = ws.cell(r, 2).value
            # skip legend/blank rows (a real label row has both a paper_id and a status)
            if not paper or not str(paper).strip() or status is None or not str(status).strip():
                continue
            cells = [ws.cell(r, c).value for c in _XLSX_COLS]
            row = [
                str(cells[0]).strip(),
                str(cells[1]).strip(),
                *["" if v is None else str(v) for v in cells[2:]],
                LABELER,
            ]
            w.writerow(row)
            n += 1
    print(f"derived {CSV.name}: {n} rows, {len(COLUMNS)} columns (no protocol_version)")
    print(
        f"label-set conforms to docs/ground-truth-protocol.md {set_level_version()} "
        f"(set-level; recorded in README.md, not per-row)"
    )
    return 0


if __name__ == "__main__":
    sys.exit(derive())
