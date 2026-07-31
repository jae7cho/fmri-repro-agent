"""Derive the canonical scored CSV from the target_space human-source labeling xlsx.

Mirrors derive_labels_csv.py (base_pipeline). Single source of truth is the **xlsx** (human-editable);
the CSV is fully REPRODUCIBLE from it, so the two cannot silently drift. Protocol version is a
**set-level** fact in target_space_README.md, NOT a per-row column.

The two "(EXAMPLE)" worked rows (oconnor, mueller) are SKIPPED: they duplicate their own real rows
(same labels), and including them would put those papers in the CSV twice. Their non-independence (the
author saw the answer in the example) is recorded in the README's blind-denominator note, not by keeping
duplicate rows.

Run:  uv run --with openpyxl python ground_truth/derive_target_space_csv.py
"""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
XLSX = HERE / "target_space_labels_v1.xlsx"
CSV = HERE / "target_space_labels_v1.csv"
README = HERE / "target_space_README.md"

LABELER = "Jae Wook Cho"
# CSV schema — deliberately NO protocol_version (set-level, see README).
COLUMNS = ["paper_id", "target_space_state", "value", "supporting_quote", "notes", "labeler"]
# xlsx Labels-sheet column order: Paper, target_space_state, Value, Supporting quote, Notes
_XLSX_COLS = (1, 2, 3, 4, 5)


def set_level_version() -> str:
    """Read the single set-level protocol version from the README's machine-readable line."""
    m = re.search(
        r"Protocol version \(set-level\): v(\d+(?:\.\d+)*)", README.read_text(encoding="utf-8")
    )
    return f"v{m.group(1)}" if m else "unknown"


def derive() -> int:
    import openpyxl

    ws = openpyxl.load_workbook(XLSX, data_only=True)["Labels"]
    n = 0
    with open(CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        w.writerow(COLUMNS)
        for r in range(2, ws.max_row + 1):
            paper = ws.cell(r, 1).value
            state = ws.cell(r, 2).value
            # skip legend/blank rows (a real label row has both a paper_id and a state)
            if not paper or not str(paper).strip() or state is None or not str(state).strip():
                continue
            # skip the two worked "(EXAMPLE)" rows — they duplicate their own real rows
            if "(EXAMPLE)" in str(paper):
                continue
            cells = [ws.cell(r, c).value for c in _XLSX_COLS]
            row = [
                str(cells[0]).strip(),
                str(cells[1]).strip(),
                *["" if v is None else str(v) for v in cells[2:]],
                LABELER,
            ]
            w.writerow(row)
            n += 1
    print(
        f"derived {CSV.name}: {n} rows, {len(COLUMNS)} columns (no protocol_version; examples skipped)"
    )
    print(
        f"label-set conforms to docs/ground-truth-protocol-target_space.md {set_level_version()} "
        f"(set-level; recorded in target_space_README.md, not per-row)"
    )
    return 0


if __name__ == "__main__":
    sys.exit(derive())
